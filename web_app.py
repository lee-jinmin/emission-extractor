import streamlit as st
import pandas as pd
import pdfplumber
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# 페이지 설정
st.set_page_config(
    page_title="배출구 데이터 추출 웹 서비스",
    page_icon="📊",
    layout="wide"
)

def detect_table_structure(table):
    """테이블 구조를 분석하여 타입을 결정"""
    if not table or len(table) < 2:
        return "unknown"
    
    # 첫 번째 행에서 구조 분석
    first_row = table[0] if table[0] else []
    second_row = table[1] if len(table) > 1 and table[1] else []
    
    # 헤더 키워드 확인
    header_text = ' '.join([str(cell) for cell in first_row + second_row if cell])
    
    if any(keyword in header_text for keyword in ['배출구', '물질명', '농도', '배출량']):
        if '최대배출기준' in header_text or '허가배출기준' in header_text:
            return "emission_standards"  # 배출기준 테이블
        else:
            return "emission_data"  # 기본 배출 데이터 테이블
    elif '허가조건' in header_text or '조건' in header_text:
        return "permit_conditions"  # 허가조건 테이블
    else:
        return "general"  # 일반 테이블

def extract_complex_table_data(table, table_type, page_num, table_idx):
    """복잡한 테이블 구조에서 데이터 추출"""
    extracted_data = []
    
    if table_type == "emission_standards":
        # 배출기준 테이블 처리
        header_row_idx = None
        
        # 헤더 찾기 (복합 헤더 고려)
        for i, row in enumerate(table):
            if row and any(cell and ('배출구' in str(cell) or '물질명' in str(cell)) for cell in row):
                header_row_idx = i
                break
        
        if header_row_idx is not None:
            # 복합 헤더 처리
            headers = []
            if header_row_idx + 1 < len(table):
                # 두 행으로 구성된 헤더인지 확인
                main_headers = table[header_row_idx]
                sub_headers = table[header_row_idx + 1] if header_row_idx + 1 < len(table) else []
                
                for i, main_header in enumerate(main_headers):
                    if main_header:
                        headers.append(str(main_header))
                    elif i < len(sub_headers) and sub_headers[i]:
                        headers.append(str(sub_headers[i]))
                    else:
                        headers.append(f"컬럼{i+1}")
                
                # 데이터 행 처리
                data_start_idx = header_row_idx + 2 if sub_headers else header_row_idx + 1
                
                for row_idx in range(data_start_idx, len(table)):
                    row = table[row_idx]
                    if row and any(cell for cell in row):
                        # 배출구 번호 확인
                        first_cell = str(row[0]) if row[0] else ""
                        if first_cell.startswith('#') or any(char.isalpha() for char in first_cell):
                            row_data = {
                                '페이지': page_num,
                                '테이블': table_idx + 1,
                                '테이블타입': table_type,
                                '원본행': row,
                                '헤더': headers
                            }
                            
                            # 각 컬럼 데이터 매핑
                            for col_idx, header in enumerate(headers):
                                if col_idx < len(row):
                                    value = str(row[col_idx]) if row[col_idx] is not None else ""
                                    row_data[header] = value
                            
                            extracted_data.append(row_data)
    
    elif table_type == "emission_data":
        # 기본 배출 데이터 테이블 처리
        for i, row in enumerate(table):
            if row and any(cell and ('배출구' in str(cell) or '#' in str(cell)) for cell in row):
                # 데이터 행 처리
                for data_row in table[i:]:
                    if data_row and len(data_row) > 0:
                        first_cell = str(data_row[0]) if data_row[0] else ""
                        if first_cell.startswith('#'):
                            row_data = {
                                '페이지': page_num,
                                '테이블': table_idx + 1,
                                '테이블타입': table_type,
                                '원본행': data_row
                            }
                            extracted_data.append(row_data)
                break
    
    return extracted_data

def extract_table_from_pdf(pdf_file, selected_outlets=['#A', '#B', '#C']):
    """PDF에서 배출구 데이터를 추출하는 함수 (개선됨)"""
    all_data = []
    page_info = []
    raw_table_data = []
    
    with pdfplumber.open(pdf_file) as pdf:
        total_pages = len(pdf.pages)
        
        for page_num, page in enumerate(pdf.pages, 1):
            # 진행률 표시
            progress = page_num / total_pages
            st.progress(progress, text=f"페이지 {page_num}/{total_pages} 처리 중...")
            
            # 테이블 추출
            tables = page.extract_tables()
            page_tables_count = 0
            page_data_count = 0
            
            if tables:
                for table_idx, table in enumerate(tables):
                    if table and len(table) > 0:
                        page_tables_count += 1
                        
                        # 테이블 구조 분석
                        table_type = detect_table_structure(table)
                        
                        # 복잡한 테이블 데이터 추출
                        complex_data = extract_complex_table_data(table, table_type, page_num, table_idx)
                        raw_table_data.extend(complex_data)
                        
                        # 기존 로직으로 기본 데이터 추출
                        header_row = None
                        for i, row in enumerate(table):
                            if row and any(cell and ('배출구' in str(cell) or '물질명' in str(cell) or '농도' in str(cell)) for cell in row):
                                header_row = i
                                break
                        
                        if header_row is not None:
                            headers = table[header_row]
                            data_rows = table[header_row + 1:]
                            
                            # 선택된 배출구 타입으로 필터링
                            for row in data_rows:
                                if row and len(row) > 0:
                                    first_cell = str(row[0]) if row[0] else ""
                                    
                                    # 선택된 배출구 타입 확인
                                    for outlet_type in selected_outlets:
                                        if first_cell.startswith(outlet_type):
                                            # 배출구 번호 추출 (더 정교한 패턴)
                                            outlet_patterns = [
                                                r'(#[A-Z]+\d*)',  # #A1, #B2 등
                                                r'(#[A-Z]+)',     # #A, #B 등
                                                r'([A-Z]+\d*)',   # A1, B2 등
                                            ]
                                            
                                            outlet_number = first_cell
                                            for pattern in outlet_patterns:
                                                match = re.match(pattern, first_cell)
                                                if match:
                                                    outlet_number = match.group(1)
                                                    break
                                            
                                            # 기본 데이터 구조
                                            row_data = {
                                                '페이지': page_num,
                                                '테이블': table_idx + 1,
                                                '테이블타입': table_type,
                                                '배출구타입': outlet_type,
                                                '배출구번호': outlet_number,
                                                '원본배출구': first_cell,
                                                '물질명': '',
                                                '농도': '',
                                                '배출량': '',
                                                '최대배출기준': '',
                                                '허가배출기준': '',
                                                '최대배출기준근거': '',
                                                '허가배출기준근거': '',
                                                '비고': '',
                                                '단위': '',
                                                '원본행': row,
                                                '헤더': headers
                                            }
                                            
                                            # 헤더와 데이터 매칭 (개선된 로직)
                                            for j, header in enumerate(headers):
                                                if j < len(row) and header:
                                                    header_str = str(header).strip().lower()
                                                    value = str(row[j]) if row[j] is not None else ""
                                                    
                                                    # 더 정교한 컬럼 매핑
                                                    if any(keyword in header_str for keyword in ['물질명', '오염물질', '항목']):
                                                        row_data['물질명'] = value
                                                    elif any(keyword in header_str for keyword in ['농도', '배출농도']):
                                                        row_data['농도'] = value
                                                    elif any(keyword in header_str for keyword in ['배출량', '연간배출량']):
                                                        row_data['배출량'] = value
                                                    elif any(keyword in header_str for keyword in ['단위']):
                                                        row_data['단위'] = value
                                                    elif '최대배출기준' in header_str:
                                                        row_data['최대배출기준'] = value
                                                    elif '허가배출기준' in header_str:
                                                        row_data['허가배출기준'] = value
                                                    elif '근거' in header_str:
                                                        if '최대' in header_str:
                                                            row_data['최대배출기준근거'] = process_emission_basis(value)
                                                        elif '허가' in header_str:
                                                            row_data['허가배출기준근거'] = process_emission_basis(value)
                                                        else:
                                                            row_data['최대배출기준근거'] = process_emission_basis(value)
                                                    elif '비고' in header_str:
                                                        # 비고에서 최대배출기준 관련 내용 제외
                                                        if value and '최대배출기준' not in value:
                                                            row_data['비고'] = value
                                                    
                                                    # 원본 헤더명으로도 저장
                                                    row_data[str(header)] = value
                                            
                                            all_data.append(row_data)
                                            page_data_count += 1
                                            break
            
            page_info.append({
                '페이지': page_num,
                '테이블수': page_tables_count,
                '추출행수': page_data_count,
                '원시데이터수': len([d for d in raw_table_data if d['페이지'] == page_num])
            })
    
    return all_data, page_info, raw_table_data

def process_emission_basis(basis_text):
    """배출기준 근거 처리 함수 (개선됨)"""
    if not basis_text:
        return ""
    
    basis_text = str(basis_text).strip()
    
    # 다양한 별표 패턴 처리
    patterns = [
        (r'별표\s*8.*?15', "별표8과15.xlsx"),
        (r'별표\s*15.*?8', "별표8과15.xlsx"),
        (r'별표.*?8', "별표8.xlsx"),
        (r'별표.*?15', "별표15.xlsx"),
    ]
    
    for pattern, replacement in patterns:
        if re.search(pattern, basis_text, re.IGNORECASE):
            return replacement
    
    # 그 외의 경우 원본 그대로 반환
    return basis_text

def validate_data_accuracy(data):
    """데이터 정확성 검증 및 불확실한 데이터 기록 (개선됨)"""
    validation_issues = []
    
    for idx, row in enumerate(data):
        issues = []
        
        # 필수 필드 체크
        required_fields = ['배출구번호', '물질명']
        for field in required_fields:
            if not row.get(field) or str(row[field]).strip() == '':
                issues.append(f"필수 필드 누락: {field}")
        
        # 배출구 번호 형식 체크
        outlet_num = row.get('배출구번호', '')
        if outlet_num and not re.match(r'^#?[A-Z]+\d*$', outlet_num):
            issues.append(f"배출구번호 형식 오류: {outlet_num}")
        
        # 숫자 필드 검증
        numeric_fields = ['농도', '배출량', '최대배출기준', '허가배출기준']
        for field in numeric_fields:
            value = row.get(field, '')
            if value and value != '-':
                # 숫자, 쉼표, 점, 하이픈, 슬래시, 공백만 허용
                clean_value = str(value).replace(',', '').replace(' ', '').replace('/', '').replace('-', '')
                if clean_value and not re.match(r'^[\d.]+$', clean_value):
                    issues.append(f"숫자 형식 오류: {field} = {value}")
        
        # 배출기준과 근거 일치성 체크
        if row.get('최대배출기준') and not row.get('최대배출기준근거'):
            issues.append("최대배출기준 근거 누락")
        
        # 테이블 타입별 추가 검증
        table_type = row.get('테이블타입', '')
        if table_type == 'emission_standards':
            if not (row.get('최대배출기준') or row.get('허가배출기준')):
                issues.append("배출기준 테이블에 기준값 없음")
        
        if issues:
            validation_issues.append({
                '행번호': idx + 1,
                '배출구': row.get('원본배출구', ''),
                '물질명': row.get('물질명', ''),
                '페이지': row.get('페이지', ''),
                '테이블타입': table_type,
                '문제점': ', '.join(issues)
            })
    
    return validation_issues

def create_standardized_excel(plan_data, review_data, validation_issues, comparison_results, unmatched_items, raw_data=None):
    """표준화된 엑셀 파일 생성 (개선됨)"""
    
    # 새 워크북 생성
    wb = openpyxl.Workbook()
    
    # 기본 시트 제거
    wb.remove(wb.active)
    
    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. 통합 데이터 시트
    ws_integrated = wb.create_sheet("통합데이터")
    
    # 헤더 설정 (확장됨)
    headers = [
        '구분', '페이지', '테이블', '테이블타입', '배출구번호', '원본배출구', '물질명', 
        '농도', '배출량', '단위', '최대배출기준', '허가배출기준', 
        '최대배출기준근거', '허가배출기준근거', '비고'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_integrated.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # 계획서 데이터 추가
    row = 2
    for data_row in plan_data:
        ws_integrated.cell(row=row, column=1, value="계획서")
        ws_integrated.cell(row=row, column=2, value=data_row.get('페이지', ''))
        ws_integrated.cell(row=row, column=3, value=data_row.get('테이블', ''))
        ws_integrated.cell(row=row, column=4, value=data_row.get('테이블타입', ''))
        ws_integrated.cell(row=row, column=5, value=data_row.get('배출구번호', ''))
        ws_integrated.cell(row=row, column=6, value=data_row.get('원본배출구', ''))
        ws_integrated.cell(row=row, column=7, value=data_row.get('물질명', ''))
        ws_integrated.cell(row=row, column=8, value=data_row.get('농도', ''))
        ws_integrated.cell(row=row, column=9, value=data_row.get('배출량', ''))
        ws_integrated.cell(row=row, column=10, value=data_row.get('단위', ''))
        ws_integrated.cell(row=row, column=11, value=data_row.get('최대배출기준', ''))
        ws_integrated.cell(row=row, column=12, value=data_row.get('허가배출기준', ''))
        ws_integrated.cell(row=row, column=13, value=data_row.get('최대배출기준근거', ''))
        ws_integrated.cell(row=row, column=14, value=data_row.get('허가배출기준근거', ''))
        ws_integrated.cell(row=row, column=15, value=data_row.get('비고', ''))
        
        # 테이블 타입별 색상 적용
        table_type = data_row.get('테이블타입', '')
        if table_type == 'emission_standards':
            type_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        elif table_type == 'emission_data':
            type_color = PatternFill(start_color="E8F0FF", end_color="E8F0FF", fill_type="solid")
        else:
            type_color = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        
        # 테두리 및 색상 적용
        for col in range(1, 16):
            cell = ws_integrated.cell(row=row, column=col)
            cell.border = border
            if col == 4:  # 테이블타입 컬럼
                cell.fill = type_color
        
        row += 1
    
    # 검토서 데이터 추가 (동일한 방식)
    for data_row in review_data:
        ws_integrated.cell(row=row, column=1, value="검토서")
        # ... (계획서와 동일한 로직)
        row += 1
    
    # 열 너비 조정
    column_widths = [8, 6, 6, 12, 12, 15, 20, 10, 10, 8, 15, 15, 20, 20, 25]
    for col, width in enumerate(column_widths, 1):
        ws_integrated.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 2. 원시 데이터 시트 (새로 추가)
    if raw_data:
        ws_raw = wb.create_sheet("원시데이터")
        
        raw_headers = ['페이지', '테이블', '테이블타입', '원본행데이터', '헤더정보']
        for col, header in enumerate(raw_headers, 1):
            cell = ws_raw.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, raw_row in enumerate(raw_data, 2):
            ws_raw.cell(row=row_idx, column=1, value=raw_row.get('페이지', ''))
            ws_raw.cell(row=row_idx, column=2, value=raw_row.get('테이블', ''))
            ws_raw.cell(row=row_idx, column=3, value=raw_row.get('테이블타입', ''))
            ws_raw.cell(row=row_idx, column=4, value=str(raw_row.get('원본행', '')))
            ws_raw.cell(row=row_idx, column=5, value=str(raw_row.get('헤더', '')))
            
            for col in range(1, 6):
                ws_raw.cell(row=row_idx, column=col).border = border
    
    # 기존 시트들 (데이터검증, 비교결과 등) 추가
    # ... (기존 코드와 동일)
    
    # 메모리에서 파일 생성
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# 메인 애플리케이션 (기존과 동일하지만 raw_data 추가)
def main():
    st.title("📊 PDF 배출구 데이터 추출 및 정리 웹 서비스")
    st.markdown("---")
    
    # 사이드바 설정
    with st.sidebar:
        st.header("⚙️ 설정")
        
        # 배출구 타입 선택
        st.subheader("배출구 타입 선택")
        outlet_types = st.multiselect(
            "추출할 배출구 타입을 선택하세요:",
            options=['#A', '#B', '#C', '#D', '#E'],
            default=['#A', '#B', '#C'],
            help="여러 타입을 선택할 수 있습니다."
        )
        
        st.markdown("---")
        st.markdown("""
        ### 📋 지원 파일 형태
        **A형태:** 기본 배출구 데이터
        **B형태:** 배출기준 포함 데이터  
        **C형태:** 복합 테이블 구조
        
        ### 🔧 처리 기능
        - 복합 헤더 구조 인식
        - 다양한 테이블 타입 자동 감지
        - 배출기준 자동 매칭
        - 데이터 정확성 검증
        - 원시 데이터 보존
        
        ### 📄 사용 방법
        1. 계획서.PDF 업로드 (필수)
        2. 검토서.PDF 업로드 (선택)
        3. 배출구 타입 선택
        4. 데이터 추출 및 정리
        5. 정리양식.xlsx 다운로드
        """)
    
    # 메인 컨텐츠
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📄 계획서.PDF 업로드")
        plan_file = st.file_uploader(
            "계획서 PDF 파일을 선택하세요",
            type=['pdf'],
            key="plan_pdf",
            help="A, B, C형태 모든 PDF 파일 지원"
        )
    
    with col2:
        st.subheader("📄 검토서.PDF 업로드 (선택사항)")
        review_file = st.file_uploader(
            "검토서 PDF 파일을 선택하세요",
            type=['pdf'],
            key="review_pdf",
            help="계획서와 비교할 검토서 PDF 파일"
        )
    
    if plan_file and outlet_types:
        st.markdown("---")
        
        if st.button("🚀 데이터 추출 및 정리 시작", type="primary"):
            with st.spinner("PDF 파일을 처리하고 있습니다..."):
                
                # 계획서 데이터 추출 (개선된 함수 사용)
                st.info("📖 계획서.PDF 데이터 추출 중...")
                plan_data, plan_page_info, plan_raw_data = extract_table_from_pdf(plan_file, outlet_types)
                
                # 검토서 데이터 추출 (있는 경우)
                review_data = []
                review_page_info = []
                review_raw_data = []
                if review_file:
                    st.info("📖 검토서.PDF 데이터 추출 중...")
                    review_data, review_page_info, review_raw_data = extract_table_from_pdf(review_file, outlet_types)
                
                # 데이터 검증
                st.info("🔍 데이터 정확성 검증 중...")
                plan_validation = validate_data_accuracy(plan_data)
                review_validation = validate_data_accuracy(review_data) if review_data else []
                all_validation_issues = plan_validation + review_validation
                
                # 결과 표시
                st.success("✅ 데이터 처리 완료!")
                
                # 통계 정보 (확장됨)
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("계획서 추출 건수", len(plan_data))
                with col2:
                    st.metric("검토서 추출 건수", len(review_data))
                with col3:
                    st.metric("원시 데이터", len(plan_raw_data) + len(review_raw_data))
                with col4:
                    st.metric("검증 이슈", len(all_validation_issues))
                
                # 테이블 타입별 통계
                if plan_data:
                    type_stats = {}
                    for data in plan_data:
                        table_type = data.get('테이블타입', 'unknown')
                        type_stats[table_type] = type_stats.get(table_type, 0) + 1
                    
                    st.subheader("📊 테이블 타입별 통계")
                    type_cols = st.columns(len(type_stats))
                    for i, (type_name, count) in enumerate(type_stats.items()):
                        with type_cols[i]:
                            st.metric(f"{type_name}", count)
                
                # 탭으로 결과 표시
                tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 통합 데이터", "🔍 원시 데이터", "⚠️ 검증 이슈", "📈 통계", "💾 다운로드"])
                
                with tab1:
                    st.subheader("통합 데이터 미리보기")
                    if plan_data:
                        plan_df = pd.DataFrame(plan_data)
                        st.dataframe(plan_df, use_container_width=True)
                
                with tab2:
                    st.subheader("원시 데이터 (디버깅용)")
                    if plan_raw_data:
                        st.write("**계획서 원시 데이터:**")
                        raw_df = pd.DataFrame(plan_raw_data)
                        st.dataframe(raw_df, use_container_width=True)
                
                with tab3:
                    st.subheader("데이터 검증 이슈")
                    if all_validation_issues:
                        st.warning(f"⚠️ {len(all_validation_issues)}개의 검증 이슈가 발견되었습니다.")
                        issues_df = pd.DataFrame(all_validation_issues)
                        st.dataframe(issues_df, use_container_width=True)
                    else:
                        st.success("✅ 검증 이슈가 없습니다.")
                
                with tab4:
                    st.subheader("처리 통계")
                    if plan_page_info:
                        st.write("**페이지별 처리 현황:**")
                        page_df = pd.DataFrame(plan_page_info)
                        st.dataframe(page_df, use_container_width=True)
                
                with tab5:
                    st.subheader("📥 정리양식.xlsx 다운로드")
                    
                    with st.spinner("엑셀 파일을 생성하고 있습니다..."):
                        excel_file = create_standardized_excel(
                            plan_data, review_data, all_validation_issues, 
                            [], [], plan_raw_data + review_raw_data
                        )
                    
                    # 파일명 생성
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"정리양식_{timestamp}.xlsx"
                    
                    st.download_button(
                        label="📥 정리양식.xlsx 다운로드",
                        data=excel_file.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )
                    
                    st.success(f"✅ {filename} 파일이 준비되었습니다!")

if __name__ == "__main__":
    main()